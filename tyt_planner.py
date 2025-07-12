import streamlit as st
import pandas as pd
import plotly.express as px
from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from groq import Groq
import numpy as np
import math

# Groq AI Client
@st.cache_resource
def init_groq_client():
    try:
        api_key = st.secrets.get("GROQ_API_KEY", "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")
        return Groq(api_key=api_key)
    except Exception as e:
        st.error(f"Groq client başlatılamadı: {str(e)}")
        return None

client = init_groq_client()

# Güncellenmiş Konu verileri (yeni soru sayılarıyla)
KONU_VERILERI ={
  "Türkçe": {
    "Paragraf": { "zorluk": "Zor", "ortalama_soru": 24, "kategori": "Dil" },
    "Cümlede Anlam": { "zorluk": "Orta", "ortalama_soru": 5, "kategori": "Dil" },
    "Sözcükte Anlam": { "zorluk": "Orta", "ortalama_soru": 3, "kategori": "Dil" },
    "Dil Bilgisi": { "zorluk": "Orta", "ortalama_soru": 4, "kategori": "Ezber" },
    "Yazım Kuralları": { "zorluk": "Kolay", "ortalama_soru": 2, "kategori": "Ezber" },
    "Noktalama İşaretleri": { "zorluk": "Kolay", "ortalama_soru": 1, "kategori": "Ezber" },
    "Ses Bilgisi": { "zorluk": "Kolay", "ortalama_soru": 1, "kategori": "Ezber" }
  },
  "Matematik": {
    "Problemler": { "zorluk": "Zor", "ortalama_soru": 12, "kategori": "Zor" },
    "Temel Kavramlar": { "zorluk": "Orta", "ortalama_soru": 3, "kategori": "Zor" },
    "Sayı Basamakları": { "zorluk": "Orta", "ortalama_soru": 2, "kategori": "Zor" },
    "Rasyonel Sayılar": { "zorluk": "Orta", "ortalama_soru": 2, "kategori": "Zor" },
    "Mutlak Değer": { "zorluk": "Orta", "ortalama_soru": 1, "kategori": "Zor" },
    "Üslü Sayılar": { "zorluk": "Orta", "ortalama_soru": 2, "kategori": "Zor" },
    "Köklü Sayılar": { "zorluk": "Orta", "ortalama_soru": 2, "kategori": "Zor" },
    "Oran-Orantı": { "zorluk": "Orta", "ortalama_soru": 2, "kategori": "Zor" },
    "Denklem Çözme": { "zorluk": "Orta", "ortalama_soru": 2, "kategori": "Zor" },
    "Kümeler": { "zorluk": "Orta", "ortalama_soru": 2, "kategori": "Zor" },
    "Fonksiyonlar": {"zorluk": "Zor", "ortalama_soru": 1, "kategori": "Zor"}
    
  },
  "Geometri": {
    "Açılar & Üçgenler": { "zorluk": "Zor", "ortalama_soru": 4, "kategori": "Zor" },
    "Katı Cisimler": { "zorluk": "Zor", "ortalama_soru": 2, "kategori": "Zor" },
    "Dikdörtgen": { "zorluk": "Orta", "ortalama_soru": 1, "kategori": "Zor" },
    "Daire": { "zorluk": "Zor", "ortalama_soru": 2, "kategori": "Zor" },
    "Analitik Geometri": { "zorluk": "Zor", "ortalama_soru": 1, "kategori": "Zor" }
  },
  "Fizik": {
    "Elektrik": { "zorluk": "Zor", "ortalama_soru": 2, "kategori": "Orta" },
    "Optik": { "zorluk": "Zor", "ortalama_soru": 2, "kategori": "Orta" },
    "Hareket": { "zorluk": "Orta", "ortalama_soru": 1, "kategori": "Orta" },
    "Isı": { "zorluk": "Orta", "ortalama_soru": 1, "kategori": "Orta" }
  },
  "Kimya": {
    "Atom": { "zorluk": "Orta", "ortalama_soru": 2, "kategori": "Orta" },
    "Tepkimeler": { "zorluk": "Zor", "ortalama_soru": 2, "kategori": "Orta" },
    "Asit-Baz": { "zorluk": "Orta", "ortalama_soru": 1, "kategori": "Orta" },
    "Organik": { "zorluk": "Zor", "ortalama_soru": 1, "kategori": "Orta" }
  },
  "Biyoloji": {
    "Hücre": { "zorluk": "Orta", "ortalama_soru": 2, "kategori": "Orta" },
    "Sınıflandırma": { "zorluk": "Kolay", "ortalama_soru": 1, "kategori": "Orta" },
    "Ekoloji": { "zorluk": "Orta", "ortalama_soru": 1, "kategori": "Orta" },
    "Genetik": { "zorluk": "Zor", "ortalama_soru": 1, "kategori": "Orta" },
    "Çevre": { "zorluk": "Orta", "ortalama_soru": 1, "kategori": "Orta" }
  },
  "Tarih": {
    "İlk ve Orta Çağ": { "zorluk": "Kolay", "ortalama_soru": 1, "kategori": "Kolay" },
    "İslamiyet'in Kabulü": { "zorluk": "Kolay", "ortalama_soru": 1, "kategori": "Kolay" },
    "Osmanlı": { "zorluk": "Orta", "ortalama_soru": 1, "kategori": "Kolay" },
    "Milli Mücadele": { "zorluk": "Orta", "ortalama_soru": 1, "kategori": "Kolay" },
    "Atatürkçülük": { "zorluk": "Kolay", "ortalama_soru": 1, "kategori": "Kolay" }
  },
  "Coğrafya": {
    "İklim": { "zorluk": "Orta", "ortalama_soru": 1, "kategori": "Kolay" },
    "Yeryüzü Şekilleri": { "zorluk": "Orta", "ortalama_soru": 1, "kategori": "Kolay" },
    "Nüfus & Yerleşme": { "zorluk": "Kolay", "ortalama_soru": 1, "kategori": "Kolay" },
    "Harita": { "zorluk": "Kolay", "ortalama_soru": 1, "kategori": "Kolay" },
    "Ekonomi": { "zorluk": "Orta", "ortalama_soru": 1, "kategori": "Kolay" }
  },
  "Felsefe": {
    "Felsefe Konusu": { "zorluk": "Zor", "ortalama_soru": 1, "kategori": "Kolay" },
    "Bilgi Felsefesi": { "zorluk": "Zor", "ortalama_soru": 1, "kategori": "Kolay" },
    "Ahlak Felsefesi": { "zorluk": "Zor", "ortalama_soru": 1, "kategori": "Kolay" },
    "Varlık Felsefesi": { "zorluk": "Zor", "ortalama_soru": 1, "kategori": "Kolay" }
  },
  "Din Kültürü ve Ahlak Bilgisi": {
    "Bilgi & İnanç": { "zorluk": "Kolay", "ortalama_soru": 1, "kategori": "Kolay" },
    "Din ve İslam": { "zorluk": "Kolay", "ortalama_soru": 1, "kategori": "Kolay" },
    "İbadet": { "zorluk": "Kolay", "ortalama_soru": 1, "kategori": "Kolay" },
    "Gençlik": { "zorluk": "Kolay", "ortalama_soru": 1, "kategori": "Kolay" }
  }
}

# Kitap ve YouTube kaynakları
KITAP_ONERILERI = {
    "Türkçe": {
        "Kolay": ["3D Türkçe Soru Bankası (Başlangıç)", "Tonguç Paragrafik", "Palme Türkçe Konu Anlatımlı", "Karekök 0 Türkçe", "Bilgiseli Türkçe"],
        "Orta": ["Limit Yayınları TYT Türkçe", "ÜçDörtBeş TYT Paragraf Soru Bankası", "Hız ve Renk Türkçe", "Apotemi Türkçe Soru Bankası", "Benim Hocam TYT Türkçe"]
    },
    "Matematik": {
        "Kolay": ["Karekök 0 Matematik", "Tonguç Kampüs TYT Matematik", "3D Matematik Konu Anlatımı", "Kolay Matematik Serisi", "Palme TYT Temel Matematik"],
        "Orta": ["ÜçDörtBeş TYT Matematik", "Limit Matematik", "Bilgiseli TYT Matematik", "Apotemi TYT Matematik", "Endemik TYT Matematik"]
    },
    "Geometri": {
        "Kolay": ["Karekök 0 Geometri", "Tonguç Kampüs TYT Geometri", "3D Geometri Konu Anlatımı", "Kolay Geometri Serisi", "Palme TYT Geometri"],
        "Orta": ["ÜçDörtBeş TYT Geometri", "Limit Geometri", "Bilgiseli TYT Geometri", "Apotemi TYT Geometri", "Endemik TYT Geometri"]
    },
    "Fizik": {
        "Kolay": ["3D TYT Fizik", "Palme Fen Bilimleri Set", "Karekök 0 Fen Serisi", "Birey A Fen Bilimleri", "Kolay Fen Bilimleri"],
        "Orta": ["ÜçDörtBeş TYT Fen Bilimleri", "Apotemi TYT Fen Set", "Hız ve Renk TYT Fizik", "Aydın TYT Fizik", "Benim Hocam Fen Seti"]
    },
    "Kimya": {
        "Kolay": ["3D TYT Kimya", "Palme Fen Bilimleri Set", "Karekök 0 Fen Serisi", "Birey A Fen Bilimleri", "Kolay Fen Bilimleri"],
        "Orta": ["ÜçDörtBeş TYT Fen Bilimleri", "Apotemi TYT Fen Set", "Hız ve Renk TYT Kimya", "Aydın TYT Kimya", "Benim Hocam Fen Seti"]
    },
    "Biyoloji": {
        "Kolay": ["3D TYT Biyoloji", "Palme Fen Bilimleri Set", "Karekök 0 Fen Serisi", "Birey A Fen Bilimleri", "Kolay Fen Bilimleri"],
        "Orta": ["ÜçDörtBeş TYT Fen Bilimleri", "Apotemi TYT Fen Set", "Hız ve Renk TYT Biyoloji", "Aydın TYT Biyoloji", "Benim Hocam Fen Seti"]
    },
    "Tarih": {
        "Kolay": ["Karekök 0 Sosyal Bilimler", "3D TYT Sosyal Bilimler", "Tonguç TYT Sosyal", "Palme Sosyal Bilimler", "Hız ve Renk TYT Sosyal"],
        "Orta": ["ÜçDörtBeş TYT Sosyal Bilimler", "Limit Yayınları TYT Sosyal", "Bilgiseli Sosyal Bilimler", "Endemik Sosyal Bilimler", "Karekök Sosyal Bilimler"]
    },
    "Coğrafya": {
        "Kolay": ["Karekök 0 Sosyal Bilimler", "3D TYT Sosyal Bilimler", "Tonguç TYT Sosyal", "Palme Sosyal Bilimler", "Hız ve Renk TYT Sosyal"],
        "Orta": ["ÜçDörtBeş TYT Sosyal Bilimler", "Limit Yayınları TYT Sosyal", "Bilgiseli Sosyal Bilimler", "Endemik Sosyal Bilimler", "Karekök Sosyal Bilimler"]
    },
    "Felsefe": {
        "Kolay": ["Karekök 0 Sosyal Bilimler", "3D TYT Sosyal Bilimler", "Tonguç TYT Sosyal", "Palme Sosyal Bilimler", "Hız ve Renk TYT Sosyal"],
        "Orta": ["ÜçDörtBeş TYT Sosyal Bilimler", "Limit Yayınları TYT Sosyal", "Bilgiseli Sosyal Bilimler", "Endemik Sosyal Bilimler", "Karekök Sosyal Bilimler"]
    },
    "Din Kültürü ve Ahlak Bilgisi": {
        "Kolay": ["Karekök 0 Sosyal Bilimler", "3D TYT Sosyal Bilimler", "Tonguç TYT Sosyal", "Palme Sosyal Bilimler", "Hız ve Renk TYT Sosyal"],
        "Orta": ["ÜçDörtBeş TYT Sosyal Bilimler", "Limit Yayınları TYT Sosyal", "Bilgiseli Sosyal Bilimler", "Endemik Sosyal Bilimler", "Karekök Sosyal Bilimler"]
    }
}

YOUTUBE_KANALLARI = {
    "Türkçe": ["Benim Hocam", "Tonguç Akademi", "Ders Vakti", "Matematik Sevdası", "Öğretmen Akademisi"],
    "Matematik": ["Tonguç Akademi", "Matematik Sevdası", "Benim Hocam", "Ders Vakti", "Matematik Dünyası"],
    "Geometri": ["Tonguç Akademi", "Matematik Sevdası", "Benim Hocam", "Ders Vakti", "Geometri Dünyası"],
    "Fizik": ["Benim Hocam", "Tonguç Akademi", "Ders Vakti", "Fizik Dünyası", "Fen Bilimleri Akademisi"],
    "Kimya": ["Benim Hocam", "Tonguç Akademi", "Ders Vakti", "Kimya Dünyası", "Fen Bilimleri Akademisi"],
    "Biyoloji": ["Benim Hocam", "Tonguç Akademi", "Ders Vakti", "Biyoloji Dünyası", "Fen Bilimleri Akademisi"],
    "Tarih": ["Benim Hocam", "Tonguç Akademi", "Ders Vakti", "Tarih Dünyası", "Sosyal Bilimler Akademisi"],
    "Coğrafya": ["Benim Hocam", "Tonguç Akademi", "Ders Vakti", "Coğrafya Dünyası", "Sosyal Bilimler Akademisi"],
    "Felsefe": ["Benim Hocam", "Tonguç Akademi", "Ders Vakti", "Felsefe Dünyası", "Sosyal Bilimler Akademisi"],
    "Din Kültürü ve Ahlak Bilgisi": ["Benim Hocam", "Tonguç Akademi", "Ders Vakti", "Din Kültürü Akademisi", "Sosyal Bilimler Akademisi"]
}

# Zaman dilimleri
ZAMAN_DILIMLERI = {
    "Zor": ["08:00-10:30", "16:00-18:00"],
    "Orta": ["10:30-12:30", "19:00-21:00"],
    "Kolay": ["13:30-15:30", "21:00-22:30"],
    "Dil": ["06:30-08:00", "22:30-23:30"],
    "Ezber": ["07:00-08:30", "22:00-23:00"]
}

# Zorluk katsayıları
ZORLUK_KATSAYILARI = {
    "Kolay": 0.5,
    "Orta": 1,
    "Zor": 2
}


# Geliştirilmiş AI Öneri Sistemi
def get_ai_suggestion(konu_analizi, gunluk_saat, gun_sayisi):
    """Geliştirilmiş ve daha detaylı AI önerisi"""
    if not client:
        return "AI hizmeti şu anda kullanılamıyor. Lütfen manuel olarak öncelikli konulara odaklanın."
    
    try:
        sorted_topics = sorted(konu_analizi.items(), key=lambda x: x[1]['oncelik_puani'], reverse=True)
        kotu_konular = sorted_topics[:8]
        orta_konular = sorted_topics[8:16] if len(sorted_topics) > 8 else []
        iyi_konular = sorted_topics[-5:]
        
        # Ders bazında analiz
        ders_analizi = {}
        for konu, info in konu_analizi.items():
            ders = info['ders']
            if ders not in ders_analizi:
                ders_analizi[ders] = {'toplam_puan': 0, 'konu_sayisi': 0, 'zayif_konular': 0}
            ders_analizi[ders]['toplam_puan'] += info['oncelik_puani']
            ders_analizi[ders]['konu_sayisi'] += 1
            if info['oncelik_puani'] > 5:
                ders_analizi[ders]['zayif_konular'] += 1
        
        for ders in ders_analizi:
            ders_analizi[ders]['ortalama'] = ders_analizi[ders]['toplam_puan'] / ders_analizi[ders]['konu_sayisi']
            ders_analizi[ders]['zayiflik_orani'] = ders_analizi[ders]['zayif_konular'] / ders_analizi[ders]['konu_sayisi']
        
        en_zayif_ders = max(ders_analizi.items(), key=lambda x: x[1]['ortalama'])
        
        # Hedef belirleme
        toplam_saat = gunluk_saat * gun_sayisi
        kritik_konu_sayisi = len([k for k, v in konu_analizi.items() if v['oncelik_puani'] > 5])
        
        prompt = f"""
        Sen TYT'de uzman bir eğitim koçusun. Türkçe cevaplamalısın sadece. Öğrencinin detaylı performans analizini yapıp, kişiselleştirilmiş {gun_sayisi} günlük strateji hazırlayacaksın.

        📊 ÖĞRENCİ PROFİLİ:
        • Toplam çalışma süresi: {toplam_saat} saat ({gun_sayisi} gün x {gunluk_saat} saat)
        • Kritik durumdaki konu sayısı: {kritik_konu_sayisi}
        • En zayıf alan: {en_zayif_ders[0]} (Risk skoru: {en_zayif_ders[1]['ortalama']:.1f})
        
        🔴 ACİL MÜDAHALE GEREKTİREN KONULAR:
        {chr(10).join([f"• {konu.split(' - ')[1]} ({konu.split(' - ')[0]}) - Risk: {info['oncelik_puani']:.1f}/10" for konu, info in kotu_konular])}
        
        🟡 GELİŞTİRİLMESİ GEREKEN KONULAR:
        {chr(10).join([f"• {konu.split(' - ')[1]} ({konu.split(' - ')[0]}) - Risk: {info['oncelik_puani']:.1f}/10" for konu, info in orta_konular])}
        
        🟢 GÜÇLÜ ALANLAR (Koruma altında):
        {chr(10).join([f"• {konu.split(' - ')[1]} ({konu.split(' - ')[0]}) - Risk: {info['oncelik_puani']:.1f}/10" for konu, info in iyi_konular])}
        
        📈 DERS BAZLI ZAYIFLIK ANALİZİ:
        {chr(10).join([f"• {ders}: %{data['zayiflik_orani']*100:.0f} zayıf konu oranı" for ders, data in ders_analizi.items()])}
        
        GÖREV: Aşağıdaki kriterlere göre {gun_sayisi} günlük DETAYLI strateji hazırla (en az 800 kelime):
        1. Kritik konular için haftalık çalışma planı (konu bazlı)
        2. Her kritik konu için özel çalışma teknikleri
        3. Zaman yönetimi stratejileri
        4. Kaynak önerileri (kitap, video, uygulama)
        5. Motivasyon teknikleri ve başarı hikayeleri
        6. Deneme sınavı takvimi
        7. Ölçme-değerlendirme yöntemleri
        8. Uyku ve beslenme düzeni önerileri
        9. Stres yönetimi teknikleri
        10. Son hafta için özel taktikler
        
        Çıktıyı başlıklar halinde düzenle ve her bölüm için en az 3-5 madde içeren detaylı açıklamalar yap.
        """
        
        chat_completion = client.chat.completions.create(
            messages=[
                {
                    "role": "system", 
                    "content": "Sen TYT'de uzman, analitik düşünen ve öğrenci psikolojisini iyi bilen bir eğitim koçusun. Veriye dayalı, kişiselleştirilmiş ve motive edici stratejiler sunuyorsun. Önerilerin en az 800 kelime olmalı ve tüm detayları kapsamalı."
                },
                {"role": "user", "content": prompt}
            ],
            model="llama3-70b-8192",
            max_tokens=4000,
            temperature=0.7
        )
        
        return chat_completion.choices[0].message.content
    except Exception as e:
        return f"AI önerisi alınırken hata oluştu: {str(e)}"


def hesapla_oncelik_puani(dogru, yanlis, bos, zorluk, ortalama_soru):
    """Geliştirilmiş öncelik puanı hesapla"""
    zorluk_katsayisi = ZORLUK_KATSAYILARI[zorluk]
    
    # Yeni formül: Yanlış ve boşları birlikte değerlendir
    puan = ((yanlis + bos) * 1.2) + (zorluk_katsayisi * 3)
    
    # Konu önem ağırlığı (logaritmik ölçek)
    onem_agirligi = np.log1p(ortalama_soru) * 3
    
    # Final puan
    oncelik_puani = puan * onem_agirligi
    
    return oncelik_puani

def analiz_et(veriler):
    """Tüm verileri analiz et"""
    analiz = {}
    for ders, konular in veriler.items():
        for konu, sonuclar in konular.items():
            if sonuclar['dogru'] + sonuclar['yanlis'] + sonuclar['bos'] > 0:
                konu_bilgi = KONU_VERILERI[ders][konu]
                
                oncelik_puani = hesapla_oncelik_puani(
                    sonuclar['dogru'], 
                    sonuclar['yanlis'], 
                    sonuclar['bos'],
                    konu_bilgi['zorluk'],
                    konu_bilgi['ortalama_soru']
                )
                
                analiz[f"{ders} - {konu}"] = {
                    'ders': ders,
                    'konu': konu,
                    'oncelik_puani': oncelik_puani,
                    'dogru': sonuclar['dogru'],
                    'yanlis': sonuclar['yanlis'],
                    'bos': sonuclar['bos'],
                    'zorluk': konu_bilgi['zorluk'],
                    'kategori': konu_bilgi['kategori'],
                    'gercek_soru': sonuclar['gercek_soru']
                }
    return analiz

def program_olustur_zaman_dilimli(analiz, baslangic_tarihi, gun_sayisi, gunluk_saat):
    """Zaman dilimli çalışma programı oluştur"""
    sorted_konular = sorted(analiz.items(), key=lambda x: x[1]['oncelik_puani'], reverse=True)
    
    program = []
    current_date = baslangic_tarihi
    
    # Kategorilere göre konuları ayır
    kategoriler = {
        "Zor": [],
        "Orta": [],
        "Kolay": [],
        "Dil": [],
        "Ezber": []
    }
    
    for konu_adi, bilgi in sorted_konular:
        kategori = bilgi['kategori']
        kategoriler[kategori].append((konu_adi, bilgi))
    
    # Her gün için program oluştur
    for gun in range(gun_sayisi):
        tarih = current_date + timedelta(days=gun)
        
        # Günlük saate göre zaman dilimlerini belirle
        if gunluk_saat <= 2:
            secilen_dilimler = ["08:00-10:30"]
        elif gunluk_saat <= 4:
            secilen_dilimler = ["08:00-10:30", "16:00-18:00"]
        elif gunluk_saat <= 6:
            secilen_dilimler = ["08:00-10:30", "10:30-12:30", "19:00-21:00"]
        else:
            secilen_dilimler = ["08:00-10:30", "10:30-12:30", "16:00-18:00", "19:00-21:00"]
        
        # Her zaman dilimine konu ata
        for zaman_dilimi in secilen_dilimler:
            # Öncelikli konu bul
            secilen_konu = None
            for kategori, konular in kategoriler.items():
                if konular:
                    secilen_konu = konular.pop(0)
                    break
            
            if secilen_konu:
                konu_adi, bilgi = secilen_konu
                
                program.append({
                    'Gün': gun + 1,
                    'Tarih': tarih.strftime('%d.%m.%Y'),
                    'Zaman': zaman_dilimi,
                    'Ders': bilgi['ders'],
                    'Konu': bilgi['konu'],
                    'Öncelik Puanı': bilgi['oncelik_puani'],
                    'Zorluk': bilgi['zorluk'],
                    'Kategori': bilgi['kategori'],
                    'Doğru': bilgi['dogru'],
                    'Yanlış': bilgi['yanlis'],
                    'Boş': bilgi['bos']
                })
    
    return program

def excel_export_professional(program_df):
    """Profesyonel Excel çıktısı"""
    output = io.BytesIO()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "TYT Çalışma Programı"
    
    # Başlık stilleri
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Kenarlık
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Koşullu biçimlendirme renkleri
    high_priority_fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
    hard_topic_fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
    
    # Başlıkları ekle
    for col, header in enumerate(program_df.columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Verileri ekle
    for row_idx, row in enumerate(dataframe_to_rows(program_df, index=False, header=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Koşullu biçimlendirme
            if col_idx == 6:  # Öncelik Puanı sütunu
                if isinstance(value, (int, float)) and value > 5:
                    cell.fill = high_priority_fill
                    cell.font = Font(bold=True, color="8B0000")
            
            if col_idx == 7:  # Zorluk sütunu
                if value == "Zor":
                    cell.fill = hard_topic_fill
                    cell.font = Font(bold=True, color="FF8C00")
    
    # Sütun genişliklerini ayarla
    column_widths = {
        'A': 8,   # Gün
        'B': 12,  # Tarih
        'C': 15,  # Zaman
        'D': 12,  # Ders
        'E': 35,  # Konu
        'F': 15,  # Öncelik Puanı
        'G': 12,  # Zorluk
        'H': 12,  # Kategori
        'I': 8,   # Doğru
        'J': 8,   # Yanlış
        'K': 8    # Boş
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Dosyayı kaydet
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()

# Yeni fonksiyon: Öğrenci performans özeti
def hesapla_performans_ozeti(veriler):
    """Öğrencinin genel performans özetini hesapla"""
    ozet = {
        'Toplam Soru': 0,
        'Toplam Doğru': 0,
        'Toplam Yanlış': 0,
        'Toplam Boş': 0,
        'Net': 0,
        'Max Net': 0,
        'Kalan Net': 0,
        'Başarı Oranı': 0
    }
    
    ders_bazli = {}
    
    for ders, konular in veriler.items():
        ders_toplam = 0
        ders_dogru = 0
        ders_yanlis = 0
        ders_bos = 0
        
        for konu, sonuclar in konular.items():
            ders_toplam += sonuclar['gercek_soru']
            ders_dogru += sonuclar['dogru']
            ders_yanlis += sonuclar['yanlis']
            ders_bos += sonuclar['bos']
            
            # Genel toplamlar
            ozet['Toplam Soru'] += sonuclar['gercek_soru']
            ozet['Toplam Doğru'] += sonuclar['dogru']
            ozet['Toplam Yanlış'] += sonuclar['yanlis']
            ozet['Toplam Boş'] += sonuclar['bos']
        
        # Ders neti (Doğru - Yanlış/4)
        ders_net = ders_dogru - (ders_yanlis / 4)
        ders_max_net = ders_toplam
        
        ders_bazli[ders] = {
            'Toplam Soru': ders_toplam,
            'Doğru': ders_dogru,
            'Yanlış': ders_yanlis,
            'Boş': ders_bos,
            'Net': ders_net,
            'Max Net': ders_max_net,
            'Kalan Net': ders_max_net - ders_net,
            'Başarı Oranı': (ders_dogru / ders_toplam * 100) if ders_toplam > 0 else 0
        }
    
    # Genel net hesaplama
    ozet['Net'] = ozet['Toplam Doğru'] - (ozet['Toplam Yanlış'] / 4)
    ozet['Max Net'] = ozet['Toplam Soru']
    ozet['Kalan Net'] = ozet['Max Net'] - ozet['Net']
    ozet['Başarı Oranı'] = (ozet['Toplam Doğru'] / ozet['Toplam Soru'] * 100) if ozet['Toplam Soru'] > 0 else 0
    
    return ozet, ders_bazli

def hesapla_ders_basari_orani(analiz_sonucu):
    """Her ders için başarı oranını hesapla"""
    ders_analizi = {}
    
    for konu_adi, bilgi in analiz_sonucu.items():
        ders = bilgi['ders']
        if ders not in ders_analizi:
            ders_analizi[ders] = {
                'toplam_puan': 0,
                'konu_sayisi': 0,
                'ortalama_puan': 0,
                'seviye': 'Kolay'
            }
        
        ders_analizi[ders]['toplam_puan'] += bilgi['oncelik_puani']
        ders_analizi[ders]['konu_sayisi'] += 1
    
    # Ortalama hesapla ve seviye belirle (YENİ ALGORİTMA)
    for ders in ders_analizi:
        ortalama = ders_analizi[ders]['toplam_puan'] / ders_analizi[ders]['konu_sayisi']
        ders_analizi[ders]['ortalama_puan'] = ortalama
        
        # Yeni seviye algoritması
        if ortalama >= 7:
            ders_analizi[ders]['seviye'] = 'Zor'
        elif ortalama >= 4:
            ders_analizi[ders]['seviye'] = 'Orta'
        else:
            ders_analizi[ders]['seviye'] = 'Kolay'
    
    return ders_analizi

def youtube_video_ara(ders_adi, konu_adi):
    """YouTube'dan video ara"""
    return [
        f"TYT {ders_adi} {konu_adi} Konu Anlatımı",
        f"TYT {ders_adi} {konu_adi} Soru Çözümü",
        f"TYT {ders_adi} {konu_adi} Tekrar Videosu",
        f"TYT {ders_adi} {konu_adi} Pratik Yöntemler"
    ]

# Streamlit arayüzü
st.set_page_config(page_title="TYT Hazırlık Uygulaması", layout="wide")

st.title("🎯 TYT Hazırlık Uygulaması")
st.markdown("---")

# Sidebar - AI Koç
with st.sidebar:
    st.header("🤖 AI Koçun")
    
    # Çalışma parametreleri
    st.subheader("📚 Çalışma Ayarları")
    gunluk_saat = st.slider("Günlük Çalışma Saati", 1, 12, 4)
    gun_sayisi = st.number_input("Kaç Gün Çalışacaksınız?", min_value=1, max_value=365, value=30)
    
    if client:
        if st.button("🔥 Kişisel Strateji Al"):
            if 'analiz_sonucu' in st.session_state:
                with st.spinner("AI senin için özel strateji hazırlıyor..."):
                    suggestion = get_ai_suggestion(st.session_state['analiz_sonucu'], gunluk_saat, gun_sayisi)
                    st.success("🎯 **Senin İçin Özel Strateji:**")
                    st.info(suggestion)
            else:
                st.warning("⚠️ Önce veri giriş yapın!")
    else:
        st.warning("AI hizmeti şu anda kullanılamıyor.")

# Ana içerik - Yeni tab ekledik
tab1, tab2, tab3, tab4, tab5 = st.tabs(["📊 Veri Giriş", "📈 Analiz", "📅 Program", "📚 Kaynaklar", "📝 Öğrenci Özeti"])

with tab1:
    st.header("Deneme Sonuçlarını Girin")
    
    if 'veriler' not in st.session_state:
        st.session_state.veriler = {}
    
    ders_gruplari = {
        "Türkçe": ["Türkçe"],
        "Matematik": ["Matematik", "Geometri"],
        "Fen Bilimleri": ["Fizik", "Kimya", "Biyoloji"],
        "Sosyal Bilimler": ["Tarih", "Coğrafya", "Felsefe", "Din Kültürü ve Ahlak Bilgisi"]
    }
    
    for grup_adi, dersler in ders_gruplari.items():
        with st.expander(f"📚 {grup_adi}", expanded=False):
            for ders in dersler:
                st.subheader(f"{ders}")
                
                if ders not in st.session_state.veriler:
                    st.session_state.veriler[ders] = {}
                
                cols = st.columns(3)
                for i, (konu, bilgi) in enumerate(KONU_VERILERI[ders].items()):
                    col_idx = i % 3
                        
                    with cols[col_idx]:
                        st.markdown(f"**{konu}**")
                        st.caption(f"Zorluk: {bilgi['zorluk']} | Ortalama: {bilgi['ortalama_soru']} soru")
                            
                        if konu not in st.session_state.veriler[ders]:
                            st.session_state.veriler[ders][konu] = {
                                'dogru': 0, 'yanlis': 0, 'bos': 0, 'gercek_soru': bilgi['ortalama_soru']
                            }
                            
                        # Gerçek soru sayısı
                        gercek_soru = st.number_input(
                            f"Denemede Bu Konudan Kaç Soru Vardı?",
                            min_value=0,
                            max_value=50,
                            key=f"{ders}_{konu}_gercek",
                            value=st.session_state.veriler[ders][konu]['gercek_soru']
                        )
                        
                        # Doğru cevaplar
                        current_dogru = st.session_state.veriler[ders][konu]['dogru']
                        if current_dogru > gercek_soru:
                            current_dogru = gercek_soru
                            
                        dogru = st.number_input(
                            f"Doğru", 
                            min_value=0, 
                            max_value=gercek_soru,
                            key=f"{ders}_{konu}_dogru", 
                            value=current_dogru
                        )
                        
                        # Yanlış cevaplar
                        current_yanlis = st.session_state.veriler[ders][konu]['yanlis']
                        max_yanlis = gercek_soru - dogru
                        
                        if current_yanlis > max_yanlis:
                            current_yanlis = max_yanlis
                            
                        yanlis = st.number_input(
                            f"Yanlış", 
                            min_value=0, 
                            max_value=max_yanlis,
                            key=f"{ders}_{konu}_yanlis",
                            value=current_yanlis
                        )
                        
                        # Boş otomatik hesapla
                        bos = max(0, gercek_soru - dogru - yanlis)
                        
                        st.text_input(
                            f"Boş (Otomatik)", 
                            value=str(bos),
                            key=f"{ders}_{konu}_bos_display",
                            disabled=True
                        )
                        
                        st.session_state.veriler[ders][konu] = {
                            'dogru': dogru,
                            'yanlis': yanlis,
                            'bos': bos,
                            'gercek_soru': gercek_soru
                        }
                        
                        # Kontrol
                        toplam = dogru + yanlis + bos
                        if toplam == gercek_soru:
                            st.success(f"✅ Toplam: {toplam}")
                        else:
                            st.error(f"❌ Toplam: {toplam}/{gercek_soru}")
with tab2:
    st.header("📊 Analiz Sonuçları")
    
    if st.button("🔍 Analiz Et"):
        analiz_sonucu = analiz_et(st.session_state.veriler)
        st.session_state.analiz_sonucu = analiz_sonucu
        
        if analiz_sonucu:
            sorted_analiz = sorted(analiz_sonucu.items(), key=lambda x: x[1]['oncelik_puani'], reverse=True)
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.subheader("🔴 Öncelikli Konular")
                for i, (konu, bilgi) in enumerate(sorted_analiz[:10]):
                    st.error(f"{i+1}. {konu} (Puan: {bilgi['oncelik_puani']:.1f})")
            
            with col2:
                st.subheader("🟢 İyi Durumda Olan Konular")
                for i, (konu, bilgi) in enumerate(sorted_analiz[-10:]):
                    st.success(f"{i+1}. {konu} (Puan: {bilgi['oncelik_puani']:.1f})")
            
            df_analiz = pd.DataFrame([
                {
                    'Konu': konu,
                    'Öncelik Puanı': bilgi['oncelik_puani'],
                    'Ders': bilgi['ders'],
                    'Zorluk': bilgi['zorluk']
                }
                for konu, bilgi in analiz_sonucu.items()
            ])
            
            fig = px.bar(df_analiz.head(20), 
                        x='Öncelik Puanı', 
                        y='Konu',
                        color='Ders',
                        title='En Öncelikli 20 Konu')
            fig.update_layout(height=600)
            st.plotly_chart(fig, use_container_width=True)
            
            # Risk haritası (Heatmap)
            st.subheader("🔥 Konu Bazlı Risk Haritası")
            
            # Heatmap için veri hazırlama
            heatmap_data = []
            for konu, bilgi in analiz_sonucu.items():
                heatmap_data.append({
                    'Ders': bilgi['ders'],
                    'Konu': bilgi['konu'],
                    'Öncelik Puanı': bilgi['oncelik_puani']
                })
            
            df_heatmap = pd.DataFrame(heatmap_data)
            
            # Pivot tablo oluşturma
            pivot_df = df_heatmap.pivot_table(
                index='Ders', 
                columns='Konu', 
                values='Öncelik Puanı', 
                aggfunc='first'
            ).fillna(0)
            
            # Heatmap oluşturma
            fig = px.imshow(
                pivot_df,
                labels=dict(x="Konu", y="Ders", color="Risk Puanı"),
                color_continuous_scale='RdYlGn_r',  # Kırmızı-Yeşil renk skalası (ters)
                title='Ders ve Konulara Göre Risk Dağılımı'
            )
            fig.update_layout(
                height=700,
                xaxis_title="Konular",
                yaxis_title="Dersler",
                coloraxis_colorbar=dict(title="Risk Puanı")
            )
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.warning("Analiz için veri bulunamadı!")

with tab3:
    st.header("📅 Çalışma Programı")
    
    if 'analiz_sonucu' in st.session_state:
        col1, col2 = st.columns(2)
        
        with col1:
            baslangic_tarihi = st.date_input("Başlangıç Tarihi", datetime.now())
        
        if st.button("📋 Zaman Dilimli Program Oluştur"):
            with st.spinner("Biyolojik saatinize uygun program hazırlanıyor..."):
                program = program_olustur_zaman_dilimli(
                    st.session_state.analiz_sonucu, 
                    baslangic_tarihi, 
                    gun_sayisi, 
                    gunluk_saat
                )
                program_df = pd.DataFrame(program)
                st.session_state.program_df = program_df
                
                # YENİ: Takvim görünümünde ders programı
                st.subheader("📅 Kişiselleştirilmiş Çalışma Takvimi")
                
                # Takvim görünümü için veriyi hazırla
                takvim_df = program_df.copy()
                takvim_df['Tarih'] = pd.to_datetime(takvim_df['Tarih'], format='%d.%m.%Y')
                takvim_df['Hafta'] = takvim_df['Tarih'].dt.isocalendar().week
                takvim_df['Gün Adı'] = takvim_df['Tarih'].dt.day_name()
                
                # Haftalara göre grupla
                haftalar = takvim_df['Hafta'].unique()
                
                for hafta in sorted(haftalar):
                    st.markdown(f"### 🗓️ Hafta {hafta}")
                    
                    # Haftanın günlerini al
                    hafta_df = takvim_df[takvim_df['Hafta'] == hafta]
                    gunler = hafta_df['Tarih'].unique()
                    
                    # Günlere göre tablolar oluştur
                    for tarih in sorted(gunler):
                        tarih_df = hafta_df[hafta_df['Tarih'] == tarih]
                        tarih_str = tarih.strftime('%d.%m.%Y')
                        
                        with st.container():
                            st.markdown(f"#### 📅 {tarih_str} ({tarih_df.iloc[0]['Gün Adı']})")
                            
                            # Her zaman dilimi için kart
                            for _, row in tarih_df.iterrows():
                                # Zorluk seviyesine göre renk
                                renk = "#FF6B6B" if row['Zorluk'] == "Zor" else "#4ECDC4" if row['Zorluk'] == "Orta" else "#FFD166"
                                
                                st.markdown(
                                    f"""
                                    <div style="
                                        background-color: {renk};
                                        border-radius: 10px;
                                        padding: 15px;
                                        margin-bottom: 15px;
                                        box-shadow: 0 4px 8px rgba(0,0,0,0.1);
                                    ">
                                        <div style="display: flex; justify-content: space-between;">
                                            <div><b>{row['Zaman']}</b></div>
                                            <div>Öncelik: {row['Öncelik Puanı']:.1f}</div>
                                        </div>
                                        <h3 style="margin: 10px 0;">{row['Ders']}</h3>
                                        <p style="margin: 0;"><b>{row['Konu']}</b></p>
                                        <p style="margin: 0; font-size: 0.9em;">Zorluk: {row['Zorluk']}</p>
                                    </div>
                                    """,
                                    unsafe_allow_html=True
                                )
                
                # İlerleme takibi
                st.subheader("📊 İlerleme Takibi")
                
                # Haftalık ilerleme grafiği
                st.markdown("#### 📈 Haftalık Konu İlerlemesi")
                haftalik_ilerleme = takvim_df.groupby('Hafta').size().reset_index(name='Konu Sayısı')
                
                fig = px.bar(haftalik_ilerleme,
                            x='Hafta',
                            y='Konu Sayısı',
                            title='Haftalık Konu İlerlemesi',
                            text='Konu Sayısı',
                            color='Konu Sayısı',
                            color_continuous_scale='Blues')
                fig.update_traces(textposition='outside')
                st.plotly_chart(fig, use_container_width=True)
                
                # Ders bazlı ilerleme
                st.markdown("#### 📚 Derslere Göre Dağılım")
                col1, col2 = st.columns(2)
                
                with col1:
                    ders_ilerleme = takvim_df.groupby('Ders').size().reset_index(name='Konu Sayısı')
                    fig = px.pie(ders_ilerleme, 
                                names='Ders', 
                                values='Konu Sayısı',
                                title='Derslere Göre Konu Dağılımı',
                                hole=0.3)
                    st.plotly_chart(fig, use_container_width=True)
                
                with col2:
                    fig = px.bar(ders_ilerleme.sort_values('Konu Sayısı', ascending=False),
                                x='Ders',
                                y='Konu Sayısı',
                                color='Ders',
                                title='Derslere Göre Konu Sayısı')
                    st.plotly_chart(fig, use_container_width=True)
    else:
        st.warning("Önce analiz yapın!")

# TAB4 İÇERİĞİ
with tab4:
    st.header("📚 Akıllı Kaynak Önerileri")
    
    if 'analiz_sonucu' in st.session_state:
        ders_basari = hesapla_ders_basari_orani(st.session_state.analiz_sonucu)
        
        st.subheader("🎯 Genel Durum Analizi")
        col1, col2, col3 = st.columns(3)
        
        zayif_dersler = [ders for ders, bilgi in ders_basari.items() if bilgi['ortalama_puan'] >= 5]
        iyi_dersler = [ders for ders, bilgi in ders_basari.items() if bilgi['ortalama_puan'] < 5]
        ortalama_risk = sum(bilgi['ortalama_puan'] for bilgi in ders_basari.values()) / len(ders_basari) if ders_basari else 0
        
        col1.metric("Zayıf Dersler", len(zayif_dersler))
        col2.metric("İyi Dersler", len(iyi_dersler))
        col3.metric("Genel Risk Skoru", f"{ortalama_risk:.1f}")
        st.markdown("---")
        
        for ders, bilgi in sorted(ders_basari.items(), key=lambda x: x[1]['ortalama_puan'], reverse=True):
            # expanded_value'yu native Python bool'una dönüştür
            expanded_value = bool(bilgi['ortalama_puan'] >= 5)
            
            with st.expander(
                f"📖 {ders} - Seviye: {bilgi['seviye']} (Risk: {bilgi['ortalama_puan']:.1f})",
                expanded=expanded_value
            ):
                st.subheader(f"📚 {ders} için Kitap Önerileri")
                seviye = bilgi['seviye']
                
                if ders in KITAP_ONERILERI and seviye in KITAP_ONERILERI[ders]:
                    kitaplar = KITAP_ONERILERI[ders][seviye]
                    cols = st.columns(2)
                    for i, kitap in enumerate(kitaplar):
                        with cols[i % 2]:
                            st.info(f"📖 {kitap}")
                else:
                    st.warning(f"{ders} için {seviye} seviye kitap önerisi bulunamadı")
                
                st.subheader(f"🎥 {ders} için YouTube Kanalları")
                if ders in YOUTUBE_KANALLARI:
                    kanallar = YOUTUBE_KANALLARI[ders]
                    cols = st.columns(3)
                    for i, kanal in enumerate(kanallar):
                        with cols[i % 3]:
                            st.success(f"📺 {kanal}")
                else:
                    st.warning(f"{ders} için YouTube kanal önerisi bulunamadı")
                
                st.subheader(f"🔍 {ders} - Zayıf Konular")
                ders_zayif_konular = [
                    (konu_adi, konu_bilgi) for konu_adi, konu_bilgi in st.session_state.analiz_sonucu.items()
                    if konu_bilgi['ders'] == ders and konu_bilgi['oncelik_puani'] >= 3
                ]
                
                if ders_zayif_konular:
                    sorted_zayif = sorted(ders_zayif_konular, key=lambda x: x[1]['oncelik_puani'], reverse=True)
                    
                    for konu_adi, konu_bilgi in sorted_zayif[:5]:
                        konu_adi_clean = konu_adi.split(' - ')[1]
                        
                        with st.container():
                            st.write(f"**{konu_adi_clean}** (Risk: {konu_bilgi['oncelik_puani']:.1f})")
                            
                            cols = st.columns(2)
                            with cols[0]:
                                st.write("🎬 **Video Önerileri:**")
                                for video in youtube_video_ara(ders, konu_adi_clean)[:3]:
                                    st.write(f"• {video}")
                            
                            with cols[1]:
                                st.write("📝 **Çalışma Önerileri:**")
                                if konu_bilgi['zorluk'] == 'Zor':
                                    st.write("• Temel kavramları tekrar edin")
                                    st.write("• Bol örnek çözün")
                                    st.write("• Günde 30 dk ayırın")
                                elif konu_bilgi['zorluk'] == 'Orta':
                                    st.write("• Soru bankası çözün")
                                    st.write("• Testler yapın")
                                    st.write("• Günde 20 dk ayırın")
                                else:
                                    st.write("• Kısa tekrarlar yapın")
                                    st.write("• Formülleri ezberleyin")
                                    st.write("• Günde 10 dk ayırın")
                            
                            st.markdown("---")
                else:
                    st.info(f"🎉 {ders} dersinde kritik zayıflık yok!")
        
        st.markdown("---")
        st.subheader("💡 Genel Strateji Önerileri")
        risk_skoru = sum(bilgi['ortalama_puan'] for bilgi in ders_basari.values()) / len(ders_basari)
        
        if risk_skoru >= 5:
            st.error("🚨 **Acil Durum Stratejisi:**")
            st.write("• Temel konulara odaklanın")
            st.write("• Günde en az 6 saat çalışın")
            st.write("• Kolay kitaplardan başlayın")
            st.write("• YouTube'dan konu anlatımları izleyin")
        elif risk_skoru >= 3:
            st.warning("⚠️ **Orta Seviye Strateji:**")
            st.write("• Zayıf konulara ağırlık verin")
            st.write("• Günde 4-5 saat çalışın")
            st.write("• Soru bankası çözmeye odaklanın")
            st.write("• Düzenli testler yapın")
        else:
            st.success("✅ **Pekiştirme Stratejisi:**")
            st.write("• Tüm konuları dengeli çalışın")
            st.write("• Günde 3-4 saat çalışın")
            st.write("• Deneme sınavlarına odaklanın")
            st.write("• Hızınızı artırmaya çalışın")
    
    else:
        st.warning("⚠️ Kaynak önerileri için önce analiz yapın!")

with tab5:
    st.header("📝 Öğrenci Genel Özeti")
    
    if 'veriler' in st.session_state and st.session_state.veriler:
        genel_ozet, ders_bazli_ozet = hesapla_performans_ozeti(st.session_state.veriler)
        
        st.subheader("📊 Genel İstatistikler")
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Toplam Soru", genel_ozet['Toplam Soru'])
        col2.metric("Toplam Doğru", genel_ozet['Toplam Doğru'])
        col3.metric("Toplam Yanlış", genel_ozet['Toplam Yanlış'])
        col4.metric("Toplam Boş", genel_ozet['Toplam Boş'])
        
        col1, col2, col3 = st.columns(3)
        col1.metric("Net", f"{genel_ozet['Net']:.2f}")
        col2.metric("Max Net", genel_ozet['Max Net'])
        col3.metric("Kalan Net", f"{genel_ozet['Kalan Net']:.2f}")
        
        st.subheader("📈 Başarı Oranı")
        basari_orani = genel_ozet['Başarı Oranı']
        st.progress(basari_orani / 100)
        st.markdown(f"**{basari_orani:.2f}%** Başarı Oranı")
        
        st.subheader("📚 Ders Bazlı Performans")
        ders_performans = []
        for ders, bilgi in ders_bazli_ozet.items():
            ders_performans.append({
                'Ders': ders, 
                'Net': bilgi['Net'], 
                'Max Net': bilgi['Max Net'], 
                'Kalan Net': bilgi['Kalan Net'], 
                'Başarı Oranı': bilgi['Başarı Oranı']
            })
        
        ders_df = pd.DataFrame(ders_performans)
        fig = px.bar(ders_df, x='Ders', y=['Net', 'Kalan Net'],
                     title='Derslere Göre Net Durumu', 
                     labels={'value': 'Net', 'variable': 'Durum'},
                     barmode='group')
        st.plotly_chart(fig, use_container_width=True)
        
        st.subheader("🔍 Detaylı Performans Tablosu")
        st.dataframe(ders_df, use_container_width=True)
        
        st.subheader("🎯 İyileştirme Alanları")
        en_cok_kalan = ders_df.sort_values('Kalan Net', ascending=False).head(3)
        for i, row in en_cok_kalan.iterrows():
            st.error(f"{row['Ders']}: {row['Kalan Net']:.2f} net kazanma potansiyeli")
    else:
        st.warning("⚠️ Önce veri girişi yapın!")

# Export butonu
if 'program_df' in st.session_state:
    st.markdown("---")
    st.subheader("📁 Dışa Aktarma")
    
    if st.button("💾 Profesyonel Excel Oluştur"):
        try:
            excel_data = excel_export_professional(st.session_state.program_df)
            st.download_button(
                label="Excel Dosyasını İndir",
                data=excel_data,
                file_name=f"tyt_program_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Excel export hatası: {str(e)}")

st.markdown("---")
st.markdown("💡 **İpucu:** Düzenli olarak deneme sonuçlarınızı güncelleyin ve programınızı yenileyin!")
